import bs4 as b
import urllib
import requests as req
import re
import openpyxl as ol
from time import sleep

# out=[]
outi=[]
ex=ol.Workbook()
ws=ex.create_sheet("total")
ws.append(["标题","热力值","链接","热点分类"])
r=req.get("https://www.zhihu.com/api/v4/creators/rank/hot?domain=0&limit=140&offset=0&period=hour")
# soup =b.BeautifulSoup(r.text,"html.parser")
tits=re.findall(r'"title":"(.*?)",',r.text)
urls=re.findall(r'"url":"(.*?)",',r.text)
hotis=re.findall(r'"score":(.*?),',r.text)
faths=re.findall(r'"topics":\[.*?\],',r.text)
for i in range(len(tits)):
    outi.append(tits[i])
    outi.append(hotis[i])
    outi.append(urls[i])
    tags=re.findall(r'"name":"(.*?)"',faths[i])
    for j in range (len(tags)):
        outi.append(tags[j])
    ws.append(outi)
    outi=[]
sleep(1)

clas=["","数码","科技","互联网","商业财经","职场","教育","法律","军事","汽车","人文社科","自然科学","工程技术","情感","心理学","两性","母婴亲子","家具","健康","艺术","音乐","设计","影视娱乐","宠物","体育电竞","运动健身","动漫游戏","美食","旅行","时尚"]
for k in range(1,30):
    ws=ex.create_sheet(clas[k])
    ws.append(["标题","热力值","链接","热点分类"])
    r=req.get("https://www.zhihu.com/api/v4/creators/rank/hot?domain=%d&limit=140&offset=0&period=hour"%(100000+k))
    # soup =b.BeautifulSoup(r.text,"html.parser")
    tits=re.findall(r'"title":"(.*?)",',r.text)
    urls=re.findall(r'"url":"(.*?)",',r.text)
    hotis=re.findall(r'"score":(.*?),',r.text)
    faths=re.findall(r'"topics":\[.*?\],',r.text)
    for i in range(len(tits)):
        outi.append(tits[i])
        outi.append(hotis[i])
        outi.append(urls[i])
        tags=re.findall(r'"name":"(.*?)"',faths[i])
        for j in range (len(tags)):
            outi.append(tags[j])
        ws.append(outi)
        outi=[]
    sleep(1)
    print(1)
#for i in range(len(tit)):
#    print (tit[i])
#    print (i)
# ws.append(out)
# print (faths)
ex.save('hot.xlsx')
# print (r.text)
# input()